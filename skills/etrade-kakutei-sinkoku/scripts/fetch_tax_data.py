"""
確定申告用 株価・為替レート取得 & 課税所得計算スクリプト
対象: ESPP / RSU (外国株) + 配当所得

使い方:
  1. 下記の「申告対象データ」セクションを実際のデータで更新する
  2. python3 fetch_tax_data.py を実行
  3. tax_summary_YYYY.xlsx / .html / .md が生成される (計算根拠・データソース付き)

データ取得元:
  - RSU/ESPP: E*TRADE → At Work → My Account → Benefit History → Download Expanded (BenefitHistory.xlsx)
  - 配当: E*TRADE → Documents → Statements → 12月分PDF の Income Summary
"""

import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ============================================================
# 申告対象データ (毎年ここを更新する)
# ============================================================

TAX_YEAR = 2025
STOCK_TICKER = "CSCO"

# RSU vesting events
# - date: vest日 (YYYY-MM-DD)
# - shares: gross株数 (税源泉前。BenefitHistoryの "Shares" 列)
# - grant_ref: グラント識別子 (メモ用)
RSU_EVENTS = [
    # BenefitHistory.xlsx から転記すること
    # {"date": "YYYY-MM-DD", "shares": N, "grant_ref": "GrantName"},
]

# ESPP purchase events
# - date: 購入日 (YYYY-MM-DD)
# - shares: 購入株数 (BenefitHistoryの "Shares Purchased" 列)
# - purchase_price: 実際の購入価格(USD) = look-back適用後 (BenefitHistoryの "Purchase Price" 列)
# - fmv: 購入日のFMV(USD) (BenefitHistoryの "FMV at Purchase" 列)
ESPP_EVENTS = [
    # BenefitHistory.xlsx から転記すること
    # {"date": "YYYY-MM-DD", "shares": N, "purchase_price": X.XXX, "fmv": YY.YY},
]

# 配当所得 (E*TRADEステートメントPDFのIncome Summary欄から転記)
DIVIDEND = {
    "qualified_usd":       0.00,   # Qualified Dividends (This Year)
    "other_usd":           0.00,   # Other Dividends (This Year)
    "us_tax_withheld_usd": 0.00,   # Tax Withholdings (This Year)
    "source_doc": "ClientStatements-XXXX-MMDDYY.pdf",  # 参照ドキュメント
    "source_page": "Income and Distribution Summary",
}

OUTPUT_BASE = f"tax_summary_{TAX_YEAR}"
OUTPUT_XLSX = f"{OUTPUT_BASE}.xlsx"
OUTPUT_HTML = f"{OUTPUT_BASE}.html"
OUTPUT_MD   = f"{OUTPUT_BASE}.md"


# ============================================================
# ヘルパー: 指定日の終値取得
# ============================================================

def get_close_price(ticker: str, date_str: str):
    """指定日の終値を取得。休日の場合は翌営業日にフォールバック。"""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    for offset in range(5):
        check = dt + timedelta(days=offset)
        end = check + timedelta(days=2)
        df = yf.download(ticker, start=check.strftime("%Y-%m-%d"),
                         end=end.strftime("%Y-%m-%d"), progress=False, auto_adjust=True)
        if not df.empty:
            close_val = df["Close"].iloc[0]
            close = float(close_val.iloc[0]) if hasattr(close_val, 'iloc') else float(close_val)
            actual_date = df.index[0].strftime("%Y-%m-%d")
            return close, actual_date
    raise ValueError(f"No price found for {ticker} near {date_str}")


# ============================================================
# 株価・為替レート取得
# ============================================================

print("=" * 60)
print(f"株価・為替レート取得中... (対象年: {TAX_YEAR})")
print("=" * 60)

all_dates = [e["date"] for e in RSU_EVENTS] + [e["date"] for e in ESPP_EVENTS]
prices = {}
for d in all_dates:
    stock_price, stock_date = get_close_price(STOCK_TICKER, d)
    usdjpy, fx_date = get_close_price("USDJPY=X", d)
    prices[d] = {
        "stock": stock_price, "stock_date": stock_date,
        "usdjpy": usdjpy, "fx_date": fx_date
    }
    print(f"  {d}: {STOCK_TICKER} ${stock_price:.2f} ({stock_date}), USD/JPY {usdjpy:.2f} ({fx_date})")

div_usdjpy, div_fx_date = get_close_price("USDJPY=X", f"{TAX_YEAR}-12-31")
print(f"  配当用 USD/JPY: {div_usdjpy:.2f} ({div_fx_date})")


# ============================================================
# RSU 課税所得計算
# ============================================================

print("\n" + "=" * 60)
print("RSU 課税所得 (給与所得)")
print("計算式: FMV(vest日株価) × 株数 × USD/JPY(vest日)")
print("=" * 60)

rsu_total_jpy = 0
rsu_rows = []
for event in RSU_EVENTS:
    d = event["date"]
    shares = event["shares"]
    stock = prices[d]["stock"]
    fx = prices[d]["usdjpy"]
    income_usd = stock * shares
    income_jpy = round(income_usd * fx)
    rsu_total_jpy += income_jpy
    rsu_rows.append({
        "【入力】vest日":           d,
        "【入力】株数(gross)":       shares,
        "【入力】グラント":           event["grant_ref"],
        "【入力】データソース":       "BenefitHistory.xlsx - Restricted Stockシート",
        "【市場】株価(USD)":         stock,
        "【市場】株価取得日":         prices[d]["stock_date"],
        "【市場】株価データソース":   f"yfinance / {STOCK_TICKER}",
        "【市場】USD/JPY":           fx,
        "【市場】為替取得日":         prices[d]["fx_date"],
        "【市場】為替データソース":   "yfinance / USDJPY=X",
        "【計算】課税所得(USD)":      round(income_usd, 2),
        "【計算】課税所得(JPY)":      income_jpy,
        "【計算】計算式":             f"{shares}株 × ${stock:.2f} × {fx:.2f}",
    })
    print(f"  {d}: {shares}株 × ${stock:.2f} × {fx:.2f} = ¥{income_jpy:,}")

print(f"\n  RSU合計課税所得: ¥{rsu_total_jpy:,}")


# ============================================================
# ESPP 課税所得計算
# ============================================================

print("\n" + "=" * 60)
print("ESPP 課税所得 (給与所得)")
print("計算式: (FMV - 購入価格) × 株数 × USD/JPY(購入日)")
print("※ 購入価格 = 85% × min(オファリング開始日FMV, 購入日FMV) [look-back規定]")
print("=" * 60)

espp_total_jpy = 0
espp_rows = []
for event in ESPP_EVENTS:
    d = event["date"]
    shares = event["shares"]
    purchase_p = event["purchase_price"]
    fmv = event["fmv"]
    fx = prices[d]["usdjpy"]
    discount_usd = (fmv - purchase_p) * shares
    income_jpy = round(discount_usd * fx)
    espp_total_jpy += income_jpy
    espp_rows.append({
        "【入力】購入日":             d,
        "【入力】株数":               shares,
        "【入力】購入価格(USD)":      purchase_p,
        "【入力】FMV(USD)":          fmv,
        "【入力】データソース":       "BenefitHistory.xlsx - ESPPシート",
        "【市場】USD/JPY":           fx,
        "【市場】為替取得日":         prices[d]["fx_date"],
        "【市場】為替データソース":   "yfinance / USDJPY=X",
        "【計算】ディスカウント/株(USD)": round(fmv - purchase_p, 3),
        "【計算】課税所得(USD)":      round(discount_usd, 2),
        "【計算】課税所得(JPY)":      income_jpy,
        "【計算】計算式":             f"({fmv:.2f}-{purchase_p:.3f}) × {shares:.3f}株 × {fx:.2f}",
    })
    print(f"  {d}: ({fmv:.2f} - {purchase_p:.3f}) × {shares:.3f}株 × {fx:.2f} = ¥{income_jpy:,}")

print(f"\n  ESPP合計課税所得: ¥{espp_total_jpy:,}")


# ============================================================
# 配当所得計算
# ============================================================

print("\n" + "=" * 60)
print("配当所得 (外国株式配当)")
print("=" * 60)

div_usd = DIVIDEND["qualified_usd"] + DIVIDEND["other_usd"]
div_jpy = round(div_usd * div_usdjpy)
tax_paid_usd = DIVIDEND["us_tax_withheld_usd"]
tax_paid_jpy = round(tax_paid_usd * div_usdjpy)

div_rows = [{
    "【入力】適格配当(Qualified, USD)":   DIVIDEND["qualified_usd"],
    "【入力】その他配当(Other, USD)":     DIVIDEND["other_usd"],
    "【入力】US源泉税(USD)":              tax_paid_usd,
    "【入力】データソース":               DIVIDEND["source_doc"],
    "【入力】参照箇所":                   DIVIDEND["source_page"],
    "【市場】USD/JPY":                    div_usdjpy,
    "【市場】為替取得日":                  div_fx_date,
    "【市場】為替データソース":           "yfinance / USDJPY=X (年末レート)",
    "【計算】配当合計(USD)":              div_usd,
    "【計算】配当合計(JPY)":              div_jpy,
    "【計算】外国税額控除(JPY)":          tax_paid_jpy,
    "【計算】計算式(配当)":              f"${div_usd:.2f} × {div_usdjpy:.2f}",
    "【計算】計算式(税額)":              f"${tax_paid_usd:.2f} × {div_usdjpy:.2f}",
}]

print(f"  配当合計(USD): ${div_usd:,.2f}  参照: {DIVIDEND['source_doc']}")
print(f"  USD/JPY: {div_usdjpy:.2f} ({div_fx_date})  参照: yfinance USDJPY=X")
print(f"  配当合計(JPY): ¥{div_jpy:,}")
print(f"  外国税額控除(JPY): ¥{tax_paid_jpy:,}")


# ============================================================
# サマリー
# ============================================================

total_salary_jpy = rsu_total_jpy + espp_total_jpy
reiwa_year = TAX_YEAR - 2018
generated_date = datetime.now().strftime("%Y-%m-%d")

print("\n" + "=" * 60)
print(f"★ 確定申告 入力サマリー (令和{reiwa_year}年分) ★")
print("=" * 60)
print(f"""
【給与所得の申告 (RSU + ESPP)】
  RSU  課税所得合計: ¥{rsu_total_jpy:>12,}
  ESPP 課税所得合計: ¥{espp_total_jpy:>12,}
  ─────────────────────────────
  給与所得加算額合計: ¥{total_salary_jpy:>12,}

【配当所得の申告 (外国株式)】
  配当収入(円換算): ¥{div_jpy:>12,}
  外国税額控除額:   ¥{tax_paid_jpy:>12,}  (米国源泉税10%)
""")

summary_rows = [
    {"申告項目": "RSU 課税所得合計",    "金額(JPY)": rsu_total_jpy,    "区分": "給与所得",  "備考": "源泉徴収票の給与所得に加算"},
    {"申告項目": "ESPP 課税所得合計",   "金額(JPY)": espp_total_jpy,   "区分": "給与所得",  "備考": "源泉徴収票の給与所得に加算"},
    {"申告項目": "給与所得 加算合計",   "金額(JPY)": total_salary_jpy, "区分": "給与所得",  "備考": "RSU + ESPP 合計"},
    {"申告項目": "配当所得",            "金額(JPY)": div_jpy,          "区分": "配当所得",  "備考": "外国株式配当(申告分離 or 総合課税)"},
    {"申告項目": "外国税額控除",        "金額(JPY)": tax_paid_jpy,     "区分": "外国税額控除", "備考": f"米国源泉税10% (日米租税条約) / ${tax_paid_usd:.2f}"},
]


# ============================================================
# Excel出力 (pandas)
# ============================================================

with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    pd.DataFrame(summary_rows).to_excel(writer,  sheet_name="申告サマリー",    index=False)
    pd.DataFrame(rsu_rows).to_excel(writer,      sheet_name="RSU_計算根拠",   index=False)
    pd.DataFrame(espp_rows).to_excel(writer,     sheet_name="ESPP_計算根拠",  index=False)
    pd.DataFrame(div_rows).to_excel(writer,      sheet_name="配当_計算根拠",  index=False)

# ============================================================
# openpyxl で書式適用
# ============================================================

wb = load_workbook(OUTPUT_XLSX)

# 色定義
COLOR_HEADER     = "1F4E79"  # 濃い青 (ヘッダー背景)
COLOR_INPUT_BG   = "DDEEFF"  # 薄い青 (入力データ行)
COLOR_MARKET_BG  = "E8F5E9"  # 薄い緑 (市場データ行)
COLOR_CALC_BG    = "FFF9C4"  # 薄い黄 (計算結果行)
COLOR_TOTAL_BG   = "FCE4EC"  # 薄い赤 (合計・申告値行)

FONT_HEADER = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FONT_INPUT  = Font(name="Arial", color="0000CD", size=10)   # 青字: 入力データ
FONT_MARKET = Font(name="Arial", color="006400", size=10)   # 緑字: 市場データ
FONT_CALC   = Font(name="Arial", color="000000", size=10)   # 黒字: 計算値
FONT_TOTAL  = Font(name="Arial", bold=True, color="8B0000", size=10)  # 赤太字: 合計

thin = Side(style="thin")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws):
    for cell in ws[1]:
        cell.font = FONT_HEADER
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def auto_col_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

def style_data_rows_by_prefix(ws):
    """列ヘッダーの【入力】【市場】【計算】プレフィックスで列ごとに色を変える"""
    headers = [ws.cell(1, c).value or "" for c in range(1, ws.max_column + 1)]
    col_fills = {}
    for i, h in enumerate(headers, 1):
        if "【入力】" in h:
            col_fills[i] = (PatternFill("solid", fgColor=COLOR_INPUT_BG), FONT_INPUT)
        elif "【市場】" in h:
            col_fills[i] = (PatternFill("solid", fgColor=COLOR_MARKET_BG), FONT_MARKET)
        elif "【計算】" in h:
            col_fills[i] = (PatternFill("solid", fgColor=COLOR_CALC_BG), FONT_CALC)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            fill_font = col_fills.get(cell.column)
            if fill_font:
                cell.fill, cell.font = fill_font
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=False)

# --- 申告サマリーシート ---
ws_sum = wb["申告サマリー"]
style_header_row(ws_sum)
for row in ws_sum.iter_rows(min_row=2):
    is_total = row[0].value and "合計" in str(row[0].value)
    for cell in row:
        cell.border = BORDER
        cell.font = FONT_TOTAL if is_total else FONT_CALC
        if is_total:
            cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
auto_col_width(ws_sum)

# --- RSU / ESPP / 配当 計算根拠シート ---
for sheet_name in ["RSU_計算根拠", "ESPP_計算根拠", "配当_計算根拠"]:
    ws = wb[sheet_name]
    style_header_row(ws)
    style_data_rows_by_prefix(ws)
    auto_col_width(ws)

# --- 凡例シートを追加 ---
ws_legend = wb.create_sheet("色の凡例")
legend_data = [
    ("色",   "区分",         "内容"),
    ("■",   "【入力】青字",  "E*TRADEから取得した生データ (BenefitHistory.xlsx / ステートメントPDF)"),
    ("■",   "【市場】緑字",  "yfinanceから取得した市場データ (株価・USD/JPY)"),
    ("■",   "【計算】黒字",  "上記を元に計算した値"),
    ("■",   "合計・申告値",  "確定申告書に入力する最終値"),
]
legend_fills = [None, COLOR_INPUT_BG, COLOR_MARKET_BG, COLOR_CALC_BG, COLOR_TOTAL_BG]
for i, (row_data, fill_color) in enumerate(zip(legend_data, legend_fills), 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_legend.cell(i, j, val)
        cell.border = BORDER
        if i == 1:
            cell.font = FONT_HEADER
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        elif fill_color and j == 1:
            cell.fill = PatternFill("solid", fgColor=fill_color)
        elif fill_color:
            cell.fill = PatternFill("solid", fgColor=fill_color)
auto_col_width(ws_legend, min_w=8, max_w=70)

# シート順を整える: 申告サマリーを先頭に
wb.move_sheet("申告サマリー", offset=-(wb.sheetnames.index("申告サマリー")))
wb.move_sheet("色の凡例", offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("色の凡例"))

wb.save(OUTPUT_XLSX)
print(f"\n{OUTPUT_XLSX} に保存しました。")
print("  シート: 申告サマリー / RSU_計算根拠 / ESPP_計算根拠 / 配当_計算根拠 / 色の凡例")
print("  色分け: 青=E*TRADEデータ, 緑=市場データ(yfinance), 黄=計算値, 赤=申告入力値")


# ============================================================
# Markdown出力
# ============================================================

def _fmt_jpy(v):
    return f"¥{v:,}"

def _fmt_usd(v):
    return f"${v:,.2f}"

# Group RSU rows by grant for cleaner presentation
rsu_by_grant = {}
for r in rsu_rows:
    g = r["【入力】グラント"]
    rsu_by_grant.setdefault(g, []).append(r)

md_lines = []
md_lines.append(f"# 確定申告データ 令和{reiwa_year}年分 ({TAX_YEAR}年)")
md_lines.append("")
md_lines.append(f"> Generated: {generated_date}")
md_lines.append(f"> Source: E*TRADE (Morgan Stanley at Work) / {STOCK_TICKER}")
md_lines.append("")
md_lines.append("---")
md_lines.append("")
md_lines.append("## 申告サマリー")
md_lines.append("")
md_lines.append("| 申告項目 | 金額 (JPY) | 区分 | 備考 |")
md_lines.append("|----------|----------:|------|------|")
for s in summary_rows:
    bold = "**" if "合計" in s["申告項目"] else ""
    md_lines.append(f"| {bold}{s['申告項目']}{bold} | {bold}{_fmt_jpy(s['金額(JPY)'])}{bold} | {bold}{s['区分']}{bold} | {bold}{s['備考']}{bold} |")
md_lines.append("")
md_lines.append("---")
md_lines.append("")

# RSU section
md_lines.append("## RSU (制限付き株式) 計算根拠")
md_lines.append("")
md_lines.append("`課税所得(JPY) = FMV(vest日株価) × 株数 × USD/JPY(vest日)`")
md_lines.append("")

for grant, rows in rsu_by_grant.items():
    md_lines.append(f"### Grant #{grant}")
    md_lines.append("")
    md_lines.append("| Vest日 | 株数 | 株価 (USD) | USD/JPY | 課税所得 (USD) | 課税所得 (JPY) |")
    md_lines.append("|--------|-----:|----------:|--------:|--------------:|--------------:|")
    subtotal_jpy = 0
    subtotal_usd = 0
    total_shares = 0
    for r in rows:
        subtotal_jpy += r["【計算】課税所得(JPY)"]
        subtotal_usd += r["【計算】課税所得(USD)"]
        total_shares += r["【入力】株数(gross)"]
        md_lines.append(f"| {r['【入力】vest日']} | {r['【入力】株数(gross)']} | {_fmt_usd(r['【市場】株価(USD)'])} | {r['【市場】USD/JPY']:.2f} | {_fmt_usd(r['【計算】課税所得(USD)'])} | {_fmt_jpy(r['【計算】課税所得(JPY)'])} |")
    md_lines.append(f"| **小計** | **{total_shares}** | | | **{_fmt_usd(subtotal_usd)}** | **{_fmt_jpy(subtotal_jpy)}** |")
    md_lines.append("")

md_lines.append(f"| | RSU合計 |")
md_lines.append(f"|---|--------:|")
md_lines.append(f"| **RSU 課税所得合計** | **{_fmt_jpy(rsu_total_jpy)}** |")
md_lines.append("")
md_lines.append("---")
md_lines.append("")

# ESPP section
md_lines.append("## ESPP (従業員株式購入計画) 計算根拠")
md_lines.append("")
md_lines.append("`課税所得(JPY) = (FMV - 購入価格) × 株数 × USD/JPY(購入日)`")
md_lines.append("")
md_lines.append("| 購入日 | 株数 | 購入価格 (USD) | FMV (USD) | ディスカウント/株 | USD/JPY | 課税所得 (USD) | 課税所得 (JPY) |")
md_lines.append("|--------|-----:|--------------:|----------:|-----------------:|--------:|--------------:|--------------:|")
total_espp_shares = 0
total_espp_usd = 0
for r in espp_rows:
    total_espp_shares += r["【入力】株数"]
    total_espp_usd += r["【計算】課税所得(USD)"]
    md_lines.append(f"| {r['【入力】購入日']} | {r['【入力】株数']:.3f} | {_fmt_usd(r['【入力】購入価格(USD)'])} | {_fmt_usd(r['【入力】FMV(USD)'])} | {_fmt_usd(r['【計算】ディスカウント/株(USD)'])} | {r['【市場】USD/JPY']:.2f} | {_fmt_usd(r['【計算】課税所得(USD)'])} | {_fmt_jpy(r['【計算】課税所得(JPY)'])} |")
md_lines.append(f"| **合計** | **{total_espp_shares:.3f}** | | | | | **{_fmt_usd(total_espp_usd)}** | **{_fmt_jpy(espp_total_jpy)}** |")
md_lines.append("")
md_lines.append("---")
md_lines.append("")

# Dividend section
md_lines.append("## 配当所得 (外国株式) 計算根拠")
md_lines.append("")
md_lines.append("| 項目 | USD | JPY | 備考 |")
md_lines.append("|------|----:|----:|------|")
md_lines.append(f"| 適格配当 (Qualified Dividends) | {_fmt_usd(DIVIDEND['qualified_usd'])} | - | CSCO配当 |")
md_lines.append(f"| その他配当 (Other Dividends) | {_fmt_usd(DIVIDEND['other_usd'])} | - | Treasury Liquidity Fund |")
md_lines.append(f"| **配当合計** | **{_fmt_usd(div_usd)}** | **{_fmt_jpy(div_jpy)}** | × {div_usdjpy:.2f} (年末レート) |")
md_lines.append(f"| US源泉税 (Tax Withholdings) | {_fmt_usd(tax_paid_usd)} | {_fmt_jpy(tax_paid_jpy)} | × {div_usdjpy:.2f} |")
md_lines.append("")
md_lines.append(f"- 為替レート: USD/JPY = {div_usdjpy:.2f} ({div_fx_date}, yfinance)")
md_lines.append(f"- 参照: {DIVIDEND['source_doc']} / {DIVIDEND['source_page']}")
md_lines.append("")
md_lines.append("---")
md_lines.append("")

# Data sources
md_lines.append("## データソース")
md_lines.append("")
md_lines.append("| データ | ソース |")
md_lines.append("|--------|--------|")
md_lines.append("| RSU vest日・株数 | BenefitHistory.xlsx (Restricted Stock sheet) |")
md_lines.append("| ESPP 購入日・株数・価格 | BenefitHistory.xlsx (ESPP sheet) |")
md_lines.append(f"| 配当・US源泉税 | {DIVIDEND['source_doc']} (12月YTD) |")
md_lines.append(f"| {STOCK_TICKER}株価 | yfinance (`{STOCK_TICKER}`) |")
md_lines.append("| USD/JPY為替レート | yfinance (`USDJPY=X`) |")
md_lines.append("")
md_lines.append("---")
md_lines.append("")

# NTA entry instructions
md_lines.append("## 確定申告書等作成コーナー 入力手順")
md_lines.append("")
md_lines.append(f"1. **給与所得 (RSU + ESPP)**: 源泉徴収票入力後に「その他給与」として **{_fmt_jpy(total_salary_jpy)}** を追加入力")
md_lines.append(f"2. **配当所得**: 「外国株式等の配当」 - 支払者: E*TRADE / Morgan Stanley, 配当金額: **{_fmt_jpy(div_jpy)}**, 国名: アメリカ合衆国")
md_lines.append(f"3. **外国税額控除**: 「外国税額控除等」 - 外国税額: **{_fmt_jpy(tax_paid_jpy)}** ({_fmt_usd(tax_paid_usd)}), 適用条約: 日米租税条約")

md_text = "\n".join(md_lines) + "\n"
Path(OUTPUT_MD).write_text(md_text, encoding="utf-8")
print(f"{OUTPUT_MD} に保存しました。")


# ============================================================
# HTML出力
# ============================================================

def _build_rsu_html_tables():
    """Build RSU HTML tables grouped by grant."""
    html = ""
    for grant, rows in rsu_by_grant.items():
        html += f'<h3>Grant #{grant}</h3>\n<table>\n<thead>\n'
        html += '<tr><th>Vest日</th><th>株数</th><th>株価 (USD)</th><th>株価取得日</th><th>USD/JPY</th><th>為替取得日</th><th>課税所得 (USD)</th><th>課税所得 (JPY)</th></tr>\n</thead>\n<tbody>\n'
        subtotal_jpy = 0
        subtotal_usd = 0
        total_shares = 0
        for r in rows:
            subtotal_jpy += r["【計算】課税所得(JPY)"]
            subtotal_usd += r["【計算】課税所得(USD)"]
            total_shares += r["【入力】株数(gross)"]
            html += f'<tr>'
            html += f'<td style="background:#DDEEFF">{r["【入力】vest日"]}</td>'
            html += f'<td class="num" style="background:#DDEEFF">{r["【入力】株数(gross)"]}</td>'
            html += f'<td class="num" style="background:#E8F5E9">{_fmt_usd(r["【市場】株価(USD)"])}</td>'
            html += f'<td style="background:#E8F5E9">{r["【市場】株価取得日"]}</td>'
            html += f'<td class="num" style="background:#E8F5E9">{r["【市場】USD/JPY"]:.2f}</td>'
            html += f'<td style="background:#E8F5E9">{r["【市場】為替取得日"]}</td>'
            html += f'<td class="num" style="background:#FFF9C4">{_fmt_usd(r["【計算】課税所得(USD)"])}</td>'
            html += f'<td class="num" style="background:#FFF9C4">{_fmt_jpy(r["【計算】課税所得(JPY)"])}</td>'
            html += '</tr>\n'
        html += f'<tr class="total-row"><td>小計</td><td class="num">{total_shares}</td><td colspan="4"></td><td class="num">{_fmt_usd(subtotal_usd)}</td><td class="num">{_fmt_jpy(subtotal_jpy)}</td></tr>\n'
        html += '</tbody>\n</table>\n'
    return html

def _build_espp_html_table():
    """Build ESPP HTML table."""
    html = '<table>\n<thead>\n'
    html += '<tr><th>購入日</th><th>株数</th><th>購入価格 (USD)</th><th>FMV (USD)</th><th>ディスカウント/株</th><th>USD/JPY</th><th>為替取得日</th><th>課税所得 (USD)</th><th>課税所得 (JPY)</th></tr>\n'
    html += '</thead>\n<tbody>\n'
    t_shares = 0
    t_usd = 0
    for r in espp_rows:
        t_shares += r["【入力】株数"]
        t_usd += r["【計算】課税所得(USD)"]
        html += '<tr>'
        html += f'<td style="background:#DDEEFF">{r["【入力】購入日"]}</td>'
        html += f'<td class="num" style="background:#DDEEFF">{r["【入力】株数"]:.3f}</td>'
        html += f'<td class="num" style="background:#DDEEFF">{_fmt_usd(r["【入力】購入価格(USD)"])}</td>'
        html += f'<td class="num" style="background:#DDEEFF">{_fmt_usd(r["【入力】FMV(USD)"])}</td>'
        html += f'<td class="num" style="background:#FFF9C4">{_fmt_usd(r["【計算】ディスカウント/株(USD)"])}</td>'
        html += f'<td class="num" style="background:#E8F5E9">{r["【市場】USD/JPY"]:.2f}</td>'
        html += f'<td style="background:#E8F5E9">{r["【市場】為替取得日"]}</td>'
        html += f'<td class="num" style="background:#FFF9C4">{_fmt_usd(r["【計算】課税所得(USD)"])}</td>'
        html += f'<td class="num" style="background:#FFF9C4">{_fmt_jpy(r["【計算】課税所得(JPY)"])}</td>'
        html += '</tr>\n'
    html += f'<tr class="total-row"><td>合計</td><td class="num">{t_shares:.3f}</td><td colspan="5"></td><td class="num">{_fmt_usd(t_usd)}</td><td class="num">{_fmt_jpy(espp_total_jpy)}</td></tr>\n'
    html += '</tbody>\n</table>\n'
    return html

html_content = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>確定申告データ 令和{reiwa_year}年分 ({TAX_YEAR}年)</title>
<style>
  :root {{ --color-header: #1F4E79; --color-input: #DDEEFF; --color-market: #E8F5E9; --color-calc: #FFF9C4; --color-total: #FCE4EC; --font-input: #0000CD; --font-market: #006400; --font-total: #8B0000; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Helvetica Neue', Arial, 'Hiragino Sans', sans-serif; color: #333; max-width: 1100px; margin: 0 auto; padding: 20px; background: #f8f9fa; }}
  h1 {{ color: var(--color-header); border-bottom: 3px solid var(--color-header); padding-bottom: 8px; margin-bottom: 6px; font-size: 1.6em; }}
  h2 {{ color: var(--color-header); margin-top: 30px; margin-bottom: 12px; font-size: 1.25em; border-left: 4px solid var(--color-header); padding-left: 10px; }}
  h3 {{ color: #555; margin-top: 18px; margin-bottom: 8px; font-size: 1.05em; }}
  .meta {{ color: #666; font-size: 0.85em; margin-bottom: 20px; }}
  table {{ border-collapse: collapse; width: 100%; margin-bottom: 16px; font-size: 0.9em; }}
  th {{ background: var(--color-header); color: #fff; padding: 8px 10px; text-align: center; font-weight: 600; white-space: nowrap; }}
  td {{ padding: 6px 10px; border: 1px solid #ccc; }}
  .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .total-row td {{ background: var(--color-total); color: var(--font-total); font-weight: bold; }}
  .summary-box {{ background: #fff; border: 2px solid var(--color-header); border-radius: 8px; padding: 20px; margin: 20px 0; }}
  .summary-box table {{ margin-bottom: 0; }}
  .formula {{ background: #f0f0f0; padding: 8px 14px; border-radius: 4px; font-family: monospace; font-size: 0.9em; margin: 8px 0 14px; display: inline-block; }}
  .note {{ background: #fffde7; border-left: 3px solid #fbc02d; padding: 10px 14px; margin: 12px 0; font-size: 0.88em; }}
  .legend {{ display: flex; gap: 16px; flex-wrap: wrap; margin: 12px 0; font-size: 0.85em; }}
  .legend-item {{ display: flex; align-items: center; gap: 5px; }}
  .legend-color {{ width: 16px; height: 16px; border: 1px solid #999; border-radius: 2px; }}
  .steps {{ background: #fff; border: 1px solid #ddd; border-radius: 6px; padding: 16px 20px; margin: 16px 0; }}
  .steps ol {{ padding-left: 20px; }}
  .steps li {{ margin-bottom: 8px; }}
  .steps strong {{ color: var(--font-total); }}
  hr {{ border: none; border-top: 1px solid #ddd; margin: 24px 0; }}
  @media print {{ body {{ background: #fff; max-width: 100%; padding: 10px; }} .summary-box {{ border-width: 1px; }} }}
</style>
</head>
<body>

<h1>確定申告データ 令和{reiwa_year}年分 ({TAX_YEAR}年)</h1>
<p class="meta">Generated: {generated_date} | Source: E*TRADE (Morgan Stanley at Work) / {STOCK_TICKER}</p>

<div class="legend">
  <div class="legend-item"><div class="legend-color" style="background:var(--color-input)"></div> E*TRADE入力データ</div>
  <div class="legend-item"><div class="legend-color" style="background:var(--color-market)"></div> 市場データ (yfinance)</div>
  <div class="legend-item"><div class="legend-color" style="background:var(--color-calc)"></div> 計算結果</div>
  <div class="legend-item"><div class="legend-color" style="background:var(--color-total)"></div> 申告入力値</div>
</div>

<div class="summary-box">
<h2 style="margin-top:0; border-left:none; padding-left:0;">申告サマリー</h2>
<table>
  <thead><tr><th>申告項目</th><th>金額 (JPY)</th><th>区分</th><th>備考</th></tr></thead>
  <tbody>
    <tr><td>RSU 課税所得合計</td><td class="num" style="background:#FFF9C4">{_fmt_jpy(rsu_total_jpy)}</td><td>給与所得</td><td>源泉徴収票の給与所得に加算</td></tr>
    <tr><td>ESPP 課税所得合計</td><td class="num" style="background:#FFF9C4">{_fmt_jpy(espp_total_jpy)}</td><td>給与所得</td><td>源泉徴収票の給与所得に加算</td></tr>
    <tr class="total-row"><td>給与所得 加算合計</td><td class="num">{_fmt_jpy(total_salary_jpy)}</td><td>給与所得</td><td>RSU + ESPP 合計</td></tr>
    <tr><td>配当所得</td><td class="num" style="background:#FFF9C4">{_fmt_jpy(div_jpy)}</td><td>配当所得</td><td>外国株式配当 (申告分離 or 総合課税)</td></tr>
    <tr class="total-row"><td>外国税額控除</td><td class="num">{_fmt_jpy(tax_paid_jpy)}</td><td>外国税額控除</td><td>米国源泉税10% (日米租税条約) / {_fmt_usd(tax_paid_usd)}</td></tr>
  </tbody>
</table>
</div>

<hr>

<h2>RSU (制限付き株式) 計算根拠</h2>
<div class="formula">課税所得(JPY) = FMV(vest日株価) &times; 株数 &times; USD/JPY(vest日)</div>

{_build_rsu_html_tables()}

<table><tr class="total-row"><td><strong>RSU 課税所得合計</strong></td><td class="num" colspan="7"><strong>{_fmt_jpy(rsu_total_jpy)}</strong></td></tr></table>

<hr>

<h2>ESPP (従業員株式購入計画) 計算根拠</h2>
<div class="formula">課税所得(JPY) = (FMV - 購入価格) &times; 株数 &times; USD/JPY(購入日)</div>

{_build_espp_html_table()}

<hr>

<h2>配当所得 (外国株式) 計算根拠</h2>
<table>
  <thead><tr><th>項目</th><th>USD</th><th>JPY</th><th>備考</th></tr></thead>
  <tbody>
    <tr><td style="background:#DDEEFF;color:#0000CD">適格配当 (Qualified Dividends)</td><td class="num" style="background:#DDEEFF">{_fmt_usd(DIVIDEND["qualified_usd"])}</td><td class="num">-</td><td>CSCO配当</td></tr>
    <tr><td style="background:#DDEEFF;color:#0000CD">その他配当 (Other Dividends)</td><td class="num" style="background:#DDEEFF">{_fmt_usd(DIVIDEND["other_usd"])}</td><td class="num">-</td><td>Treasury Liquidity Fund</td></tr>
    <tr class="total-row"><td>配当合計</td><td class="num">{_fmt_usd(div_usd)}</td><td class="num">{_fmt_jpy(div_jpy)}</td><td>&times; {div_usdjpy:.2f} (年末レート)</td></tr>
    <tr><td style="background:#DDEEFF;color:#0000CD">US源泉税 (Tax Withholdings)</td><td class="num" style="background:#DDEEFF">{_fmt_usd(tax_paid_usd)}</td><td class="num" style="background:#FFF9C4">{_fmt_jpy(tax_paid_jpy)}</td><td>&times; {div_usdjpy:.2f}</td></tr>
  </tbody>
</table>
<div class="note">
  <strong>為替レート:</strong> USD/JPY = {div_usdjpy:.2f} ({div_fx_date}, yfinance USDJPY=X)<br>
  <strong>参照:</strong> {DIVIDEND["source_doc"]} / {DIVIDEND["source_page"]}
</div>

<hr>

<h2>データソース</h2>
<table>
  <thead><tr><th>データ</th><th>ソース</th></tr></thead>
  <tbody>
    <tr><td>RSU vest日・株数</td><td>BenefitHistory.xlsx (Restricted Stock sheet)</td></tr>
    <tr><td>ESPP 購入日・株数・価格</td><td>BenefitHistory.xlsx (ESPP sheet)</td></tr>
    <tr><td>配当・US源泉税</td><td>{DIVIDEND["source_doc"]} (12月YTD)</td></tr>
    <tr><td>{STOCK_TICKER}株価</td><td>yfinance ({STOCK_TICKER})</td></tr>
    <tr><td>USD/JPY為替レート</td><td>yfinance (USDJPY=X)</td></tr>
  </tbody>
</table>

<hr>

<h2>確定申告書等作成コーナー 入力手順</h2>
<div class="steps">
<ol>
  <li><strong>給与所得 (RSU + ESPP):</strong> 源泉徴収票入力後に「その他給与」として <strong>{_fmt_jpy(total_salary_jpy)}</strong> を追加入力</li>
  <li><strong>配当所得:</strong> 「外国株式等の配当」 &mdash; 支払者: E*TRADE / Morgan Stanley, 配当金額: <strong>{_fmt_jpy(div_jpy)}</strong>, 国名: アメリカ合衆国</li>
  <li><strong>外国税額控除:</strong> 「外国税額控除等」 &mdash; 外国税額: <strong>{_fmt_jpy(tax_paid_jpy)}</strong> ({_fmt_usd(tax_paid_usd)}), 適用条約: 日米租税条約</li>
</ol>
</div>

</body>
</html>"""

Path(OUTPUT_HTML).write_text(html_content, encoding="utf-8")
print(f"{OUTPUT_HTML} に保存しました。")

print(f"\n出力ファイル一覧:")
print(f"  {OUTPUT_XLSX} (Excel - 色分け付き計算根拠)")
print(f"  {OUTPUT_MD}   (Markdown)")
print(f"  {OUTPUT_HTML} (HTML - 印刷対応・色分け凡例付き)")
